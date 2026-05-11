"""Excel timetable export using openpyxl."""
from __future__ import annotations
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
HEADER_FILL = PatternFill("solid", fgColor="1E40AF")
ALT_FILL = PatternFill("solid", fgColor="EFF6FF")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def generate_timetable_excel(timetable: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    # Meta rows
    ws.append([f"Timetable: {timetable.get('name', '')}"])
    ws.append([f"Semester: {timetable.get('semester')} | Year: {timetable.get('academic_year')}"])
    ws.append([f"Fitness Score: {timetable.get('fitness_score', 0):.2%} | "
               f"Generated in: {timetable.get('generation_time_seconds', 0):.1f}s"])
    ws.append([])

    # Header row
    ws.append(["Time Slot"] + DAYS)
    header_row = ws.max_row
    for col in range(1, 7):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Build slot grid
    entries = timetable.get("entries", [])
    slot_map: dict[tuple, dict] = {}
    for entry in entries:
        slot = entry.get("time_slot", {})
        day = slot.get("day", "")
        time = f"{slot.get('start_time', '')}–{slot.get('end_time', '')}"
        key = (time, slot.get("slot_index", 0))
        slot_map.setdefault(key, {})
        slot_map[key][day] = entry

    time_keys = sorted(slot_map.keys(), key=lambda k: k[1])
    for row_idx, ((time_label, _), day_entries) in enumerate(
        zip(time_keys, [slot_map[k] for k in time_keys])
    ):
        row = [time_label]
        for day in DAYS:
            entry = day_entries.get(day)
            if entry:
                course = entry.get("course", {})
                room = entry.get("room", {})
                lecturer = entry.get("lecturer", {})
                row.append(f"{course.get('code', '')} | {course.get('name', '')}\n"
                           f"Room: {room.get('code', '')} | {lecturer.get('name', '')}")
            else:
                row.append("")
        ws.append(row)

        # Style
        data_row = ws.max_row
        fill = ALT_FILL if row_idx % 2 else PatternFill()
        for col in range(1, 7):
            cell = ws.cell(row=data_row, column=col)
            if col > 1 and fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 16
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 28

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
